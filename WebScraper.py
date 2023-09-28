import json
import scrapy
import Environment
from AzureKeyPhrase import azureService, score
from urllib.parse import urlencode
from SearchData import cleanUrl, searchData
from progress.bar import ChargingBar
import time, random
from scrapy.crawler import CrawlerProcess
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from datetime import datetime
import scrapy.utils.misc 
import scrapy.core.scraper

def warn_on_generator_with_return_value_stub(spider, callable):
    pass

scrapy.utils.misc.warn_on_generator_with_return_value = warn_on_generator_with_return_value_stub
scrapy.core.scraper.warn_on_generator_with_return_value = warn_on_generator_with_return_value_stub

class test (scrapy.Spider):
    name = "test"
    start_url = ""
    proxy_url = ""
    number = 0
    max_items = 5
    azure = False
    vector =[]
    excel_wb = openpyxl.Workbook()
    sheet = excel_wb.active 
    def start_requests(self):
        yield scrapy.Request(self.proxy_url, callback=self.startGoogle)


    def startGoogle(self,response):
        resultJson=json.loads(response.text)
        for result in resultJson["organic_results"]:
            self.vector = cleanUrl(result["link"], self.vector, self.max_items)
        if len(self.vector) < self.max_items:
            self.number +=10
            createUrl(self.number)
            yield scrapy.Request(self.proxy_url, callback=self.startGoogle)
        else:
            for link in self.vector:
                yield scrapy.Request(link, callback = self.parse)
        
    def parse(self, response):
        
        selectors=[ "a::text",
                    "h3::text",
                    "p::text",
                    "h2::text",
                    "strong::text",
                    "span::text",
                    "div::text", 
                    "li::text",
                    "a::attr(href)"]

        data = {"name": [],
                "url": [response.request.url],
                "email": [],
                "phone":[],
                "address":[],
                "score":[],
                "social media":[]}

        document= documentAzure(response.css("p::text").getall())
        data["score"]=[azureService(document, self.azure)]

        find_Data_Name= response.css("title::text").getall()
        for name in find_Data_Name:
            data["name"].append(name.strip())
        for i in selectors:
            find_Data = response.css(i).getall()
            for i in find_Data:
                dato, key = searchData(i)
                if dato != "" and dato not in data[key]:
                    data[key].append(dato.strip())
        print(data["name"])
        self.write_sheet(data)
        progressBar()

    def write_sheet(self, coll):
        sheet_name = buildsheetname(coll['name'][0])
        self.excel_wb.create_sheet(sheet_name)
        self.sheet = self.excel_wb[sheet_name]
        self.excel_wb.active = self.sheet
        self.sheet.cell(row = 1, column = 1).value = coll['name'][0].upper()
        self.sheet.cell(row = 1, column = 1).font = Font(size = 18 , name = 'Arabic Typesetting', bold = True)
        self.sheet.cell(row = 1, column = 1).alignment = Alignment(horizontal="center", vertical="center")
        row = 2
        col = 1
        for a_row, a_col in coll.items():
            self.sheet.column_dimensions[get_column_letter(col)].width = width_columns[col-1] 
            row = 2
            if a_row == "name":
                continue
            self.sheet.cell(row = row, column = col).value = a_row.capitalize()
            self.sheet.cell(row = row, column = col).font = Font(size = 14 , name = 'Bell MT', bold = True)
            self.sheet.cell(row = row, column = col).alignment = Alignment(horizontal="center", vertical="center")
            for a_data in a_col:
                row += 1
                self.sheet.cell(row = row, column = col).value = a_data
            col += 1
        self.sheet.merge_cells('A1:F1')
        filename = "Scrapy_Result{}.xlsx".format(date.strftime("_%Y_%m_%d_%H_%M_%S"))
        self.excel_wb.save(filename)


def progressBar():
    print("")
    bar2 = ChargingBar('Finding:', max=100)
    for num in range(100):
        time.sleep(random.uniform(0, 0.01))
        bar2.next()
    bar2.finish()
    print("")

def buildsheetname(text):
    if len(text) > 30:
        return text[0:30]
    return text

def createUrl(number):
    payLoad= {"api_key": Environment.apiKey, "url": test.start_url + str(number), "autoparse":"auto"}
    test.proxy_url= Environment.apiEndpoint+ urlencode(payLoad)


def geturl(url, items, azure):
    test.start_url = url
    test.max_items = items 
    test.azure= azure
    createUrl(test.number)
    process = CrawlerProcess({ 
    "LOG_LEVEL": "ERROR",
    })
    process.crawl(test)
    process.start()  

def documentAzure(docuemnt):
    newDocument=[""]
    pos = 0
    for phrase in docuemnt:
        if len(newDocument[pos] + phrase)<5000:
            newDocument[pos] = newDocument[pos] + " " + phrase
        else:
            if pos == 9:
                return newDocument
            newDocument.append(phrase)
            pos += 1
    return newDocument

width_columns = [70, 50, 20, 30, 10, 30] 
date=datetime.now()
        